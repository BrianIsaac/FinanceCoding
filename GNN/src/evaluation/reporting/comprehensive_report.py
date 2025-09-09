"""
Comprehensive performance report generation and export framework.

This module integrates all reporting components to generate complete institutional-grade
reports with executive summaries, detailed analysis, feasibility assessments, strategic
recommendations, and multi-format export capabilities.
"""

from __future__ import annotations

import warnings
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd

try:
    import matplotlib.pyplot as plt
    import seaborn as sns

    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False
    warnings.warn("Matplotlib/Seaborn not available. Static plotting disabled.", stacklevel=2)

try:
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots

    HAS_PLOTLY = True
except ImportError:
    HAS_PLOTLY = False
    warnings.warn("Plotly not available. Interactive plotting disabled.", stacklevel=2)

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False
    warnings.warn("ReportLab not available. PDF generation disabled.", stacklevel=2)

try:
    import openpyxl
    from openpyxl.styles import Border, Font, PatternFill, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    warnings.warn("OpenPyXL not available. Excel export disabled.", stacklevel=2)

from src.evaluation.reporting.executive import ExecutiveSummaryGenerator
from src.evaluation.reporting.feasibility import ImplementationFeasibilityAssessor
from src.evaluation.reporting.interactive import InteractiveDashboard
from src.evaluation.reporting.performance_analysis import ComprehensivePerformanceAnalyzer
from src.evaluation.reporting.strategy import StrategyRecommendationEngine


@dataclass
class ReportConfig:
    """Configuration for comprehensive report generation."""

    output_directory: str = "reports"
    report_title: str = "ML-Enhanced Portfolio Construction Analysis"
    institution_name: str = "Investment Management Firm"

    # Export formats
    generate_pdf: bool = True
    generate_excel: bool = True
    generate_html: bool = True
    generate_interactive: bool = True

    # Report sections
    include_executive_summary: bool = True
    include_performance_analysis: bool = True
    include_feasibility_assessment: bool = True
    include_strategic_recommendations: bool = True
    include_appendices: bool = True

    # Styling
    color_scheme: str = "professional"  # professional, modern, classic
    font_family: str = "Arial"
    logo_path: str = None


class ComprehensiveReportGenerator:
    """
    Comprehensive report generator integrating all analysis components.

    Generates institutional-grade performance reports with executive summaries,
    detailed analysis, strategic recommendations, and multi-format export.
    """

    def __init__(self, config: ReportConfig = None):
        """
        Initialize comprehensive report generator.

        Args:
            config: Report generation configuration
        """
        self.config = config or ReportConfig()

        # Initialize component generators
        self.executive_generator = ExecutiveSummaryGenerator()
        self.performance_analyzer = ComprehensivePerformanceAnalyzer()
        self.feasibility_assessor = ImplementationFeasibilityAssessor()
        self.strategy_engine = StrategyRecommendationEngine()

        if HAS_PLOTLY:
            self.interactive_dashboard = InteractiveDashboard()

    def generate_comprehensive_report(
        self,
        performance_results: dict[str, pd.DataFrame],
        operational_metrics: dict[str, dict[str, Any]],
        statistical_results: dict[str, Any],
        model_specifications: dict[str, dict[str, Any]],
        institutional_constraints: dict[str, Any] = None,
        output_path: Path = None,
    ) -> dict[str, Any]:
        """
        Generate comprehensive performance report with all components.

        Args:
            performance_results: Model performance results by approach
            operational_metrics: Operational efficiency metrics
            statistical_results: Statistical significance test results
            model_specifications: Technical model specifications
            institutional_constraints: Institution-specific constraints
            output_path: Optional custom output path

        Returns:
            Dictionary containing all report components and metadata
        """
        # Set up output directory
        if output_path is None:
            output_path = Path(self.config.output_directory)
        output_path.mkdir(parents=True, exist_ok=True)

        report_components = {}

        # Generate Executive Summary
        if self.config.include_executive_summary:
            executive_data = self.executive_generator.generate_executive_dashboard(
                performance_results,
                operational_metrics,
                statistical_results,
                output_path / "executive",
            )
            report_components["executive_summary"] = executive_data

        # Generate Detailed Performance Analysis
        if self.config.include_performance_analysis:
            performance_tables = (
                self.performance_analyzer.generate_comprehensive_performance_tables(
                    performance_results, statistical_results, output_dir=output_path / "performance"
                )
            )
            report_components["performance_analysis"] = performance_tables

        # Generate Feasibility Assessment
        if self.config.include_feasibility_assessment:
            computational_analysis = self.feasibility_assessor.assess_computational_requirements(
                model_specifications, operational_metrics
            )
            operational_analysis = self.feasibility_assessor.assess_operational_complexity(
                model_specifications, operational_metrics
            )
            tco_analysis = self.feasibility_assessor.calculate_total_cost_ownership(
                computational_analysis, operational_analysis
            )
            timeline_analysis = self.feasibility_assessor.assess_implementation_timeline(
                model_specifications
            )

            feasibility_summary = self.feasibility_assessor.create_feasibility_summary_report(
                computational_analysis,
                operational_analysis,
                tco_analysis,
                timeline_analysis,
                output_path / "feasibility_summary.txt",
            )

            report_components["feasibility_assessment"] = {
                "computational_requirements": computational_analysis,
                "operational_complexity": operational_analysis,
                "tco_analysis": tco_analysis,
                "timeline_analysis": timeline_analysis,
                "feasibility_summary": feasibility_summary,
            }

        # Generate Strategic Recommendations
        if self.config.include_strategic_recommendations:
            if institutional_constraints is None:
                institutional_constraints = self._default_institutional_constraints()

            # Create combined feasibility results for strategy engine
            if "feasibility_assessment" in report_components:
                feasibility_results = report_components["feasibility_assessment"]["tco_analysis"]
            else:
                feasibility_results = pd.DataFrame()

            # Generate decision framework
            performance_summary = report_components.get("executive_summary", {}).get(
                "rankings", pd.DataFrame()
            )
            if not performance_summary.empty and not feasibility_results.empty:
                decision_framework = self.strategy_engine.create_decision_tree_framework(
                    performance_summary, feasibility_results, institutional_constraints
                )

                # Generate implementation roadmap
                recommended_approaches = decision_framework["implementation_priority"][:3]
                timeline_data = report_components["feasibility_assessment"]["timeline_analysis"]
                implementation_roadmap = self.strategy_engine.generate_implementation_roadmap(
                    recommended_approaches, timeline_data, institutional_constraints
                )

                report_components["strategic_recommendations"] = {
                    "decision_framework": decision_framework,
                    "implementation_roadmap": implementation_roadmap,
                }

        # Generate reports in requested formats
        report_files = {}

        if self.config.generate_pdf and HAS_REPORTLAB:
            pdf_path = self._generate_pdf_report(
                report_components, output_path / "comprehensive_report.pdf"
            )
            report_files["pdf"] = pdf_path

        if self.config.generate_excel and HAS_OPENPYXL:
            excel_path = self._generate_excel_report(
                report_components, output_path / "comprehensive_report.xlsx"
            )
            report_files["excel"] = excel_path

        if self.config.generate_html:
            html_path = self._generate_html_report(
                report_components, output_path / "comprehensive_report.html"
            )
            report_files["html"] = html_path

        if self.config.generate_interactive and HAS_PLOTLY:
            dashboard_path = self._generate_interactive_dashboard(
                report_components, output_path / "interactive_dashboard.html"
            )
            report_files["interactive"] = dashboard_path

        # Generate summary metadata
        report_metadata = {
            "generation_timestamp": pd.Timestamp.now().isoformat(),
            "models_analyzed": len(performance_results),
            "report_sections": list(report_components.keys()),
            "output_formats": list(report_files.keys()),
            "output_directory": str(output_path),
        }

        return {
            "components": report_components,
            "files": report_files,
            "metadata": report_metadata,
        }

    def _default_institutional_constraints(self) -> dict[str, Any]:
        """Default institutional constraints for analysis."""
        return {
            "risk_tolerance": "moderate",
            "aum_millions": 500,
            "regulatory_complexity": "standard",
            "computational_budget": "medium",
            "preferred_timeline_months": 6,
            "team_experience_level": "medium",
            "existing_infrastructure": "partial",
        }

    def _generate_pdf_report(self, components: dict[str, Any], output_path: Path) -> Path:
        """Generate PDF report using ReportLab."""
        if not HAS_REPORTLAB:
            return None

        doc = SimpleDocTemplate(str(output_path), pagesize=letter)
        styles = getSampleStyleSheet()
        story = []

        # Title
        title = Paragraph(self.config.report_title, styles["Title"])
        story.append(title)
        story.append(Spacer(1, 12))

        # Executive Summary
        if "executive_summary" in components:
            story.append(Paragraph("Executive Summary", styles["Heading1"]))
            summary_stats = components["executive_summary"].get("summary_statistics", {})

            summary_text = f"""
            Total Models Evaluated: {summary_stats.get('total_models_evaluated', 'N/A')}
            Top Performing Model: {summary_stats.get('top_performer', 'N/A')}
            Best Sharpe Ratio: {summary_stats.get('best_sharpe_ratio', 'N/A'):.4f}
            Recommended Model: {summary_stats.get('recommended_model', 'N/A')}
            """
            story.append(Paragraph(summary_text, styles["Normal"]))
            story.append(Spacer(1, 12))

        # Performance Analysis
        if "performance_analysis" in components:
            story.append(Paragraph("Performance Analysis", styles["Heading1"]))

            # Add performance summary table
            perf_summary = components["performance_analysis"].get(
                "performance_summary", pd.DataFrame()
            )
            if not perf_summary.empty:
                # Convert DataFrame to table data
                table_data = [perf_summary.columns.tolist()]
                for _, row in perf_summary.head(10).iterrows():  # Limit rows for PDF
                    table_data.append(row.astype(str).tolist())

                table = Table(table_data)
                table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, 0), 10),
                            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                            ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ]
                    )
                )
                story.append(table)
                story.append(Spacer(1, 12))

        # Strategic Recommendations
        if "strategic_recommendations" in components:
            story.append(Paragraph("Strategic Recommendations", styles["Heading1"]))
            decision_framework = components["strategic_recommendations"].get(
                "decision_framework", {}
            )

            rec_text = f"""
            Primary Recommendation: {decision_framework.get('Primary Decision', 'N/A')}
            Rationale: {decision_framework.get('Rationale', 'N/A')}
            """
            story.append(Paragraph(rec_text, styles["Normal"]))

        # Build PDF
        doc.build(story)
        return output_path

    def _generate_excel_report(self, components: dict[str, Any], output_path: Path) -> Path:
        """Generate Excel report with multiple sheets."""
        if not HAS_OPENPYXL:
            return None

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Executive Summary Sheet
            if "executive_summary" in components:
                rankings = components["executive_summary"].get("rankings", pd.DataFrame())
                if not rankings.empty:
                    rankings.to_excel(writer, sheet_name="Executive Summary", index=False)

            # Performance Analysis Sheets
            if "performance_analysis" in components:
                perf_analysis = components["performance_analysis"]
                for table_name, table_df in perf_analysis.items():
                    if isinstance(table_df, pd.DataFrame) and not table_df.empty:
                        sheet_name = table_name.replace("_", " ").title()[
                            :31
                        ]  # Excel sheet name limit
                        table_df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Feasibility Assessment Sheets
            if "feasibility_assessment" in components:
                feas_assessment = components["feasibility_assessment"]
                for analysis_name, analysis_df in feas_assessment.items():
                    if isinstance(analysis_df, pd.DataFrame) and not analysis_df.empty:
                        sheet_name = analysis_name.replace("_", " ").title()[:31]
                        analysis_df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Strategic Recommendations Sheet
            if "strategic_recommendations" in components:
                strategy_data = components["strategic_recommendations"]
                decision_df = pd.DataFrame([strategy_data.get("decision_framework", {})])
                if not decision_df.empty:
                    decision_df.to_excel(
                        writer, sheet_name="Strategic Recommendations", index=False
                    )

        # Format Excel file
        self._format_excel_report(output_path)
        return output_path

    def _format_excel_report(self, excel_path: Path) -> None:
        """Apply formatting to Excel report."""
        if not HAS_OPENPYXL:
            return

        workbook = openpyxl.load_workbook(excel_path)

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Format each sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Format headers
            for cell in sheet[1]:  # First row
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border

            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(excel_path)

    def _generate_html_report(self, components: dict[str, Any], output_path: Path) -> Path:
        """Generate HTML report with embedded styling."""
        html_content = self._create_html_template()

        # Add executive summary
        if "executive_summary" in components:
            html_content += "<h2>Executive Summary</h2>\n"
            rankings = components["executive_summary"].get("rankings", pd.DataFrame())
            if not rankings.empty:
                html_content += rankings.to_html(classes="table table-striped", escape=False)

        # Add performance analysis
        if "performance_analysis" in components:
            html_content += "<h2>Performance Analysis</h2>\n"
            perf_analysis = components["performance_analysis"]
            for table_name, table_df in perf_analysis.items():
                if isinstance(table_df, pd.DataFrame) and not table_df.empty:
                    html_content += f"<h3>{table_name.replace('_', ' ').title()}</h3>\n"
                    html_content += table_df.to_html(classes="table table-striped", escape=False)

        # Close HTML
        html_content += "</div></body></html>"

        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html_content)

        return output_path

    def _create_html_template(self) -> str:
        """Create HTML template with styling."""
        return f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{self.config.report_title}</title>
    <style>
        body {{
            font-family: {self.config.font_family}, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f8f9fa;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #2c3e50;
            border-bottom: 3px solid #3498db;
            padding-bottom: 10px;
        }}
        h2 {{
            color: #34495e;
            margin-top: 30px;
            margin-bottom: 15px;
        }}
        h3 {{
            color: #7f8c8d;
            margin-top: 20px;
            margin-bottom: 10px;
        }}
        .table {{
            width: 100%;
            margin-bottom: 20px;
            border-collapse: collapse;
        }}
        .table th, .table td {{
            padding: 8px;
            text-align: left;
            border: 1px solid #dee2e6;
        }}
        .table th {{
            background-color: #f8f9fa;
            font-weight: bold;
        }}
        .table-striped tbody tr:nth-child(odd) {{
            background-color: rgba(0,0,0,.05);
        }}
    </style>
</head>
<body>
<div class="container">
    <h1>{self.config.report_title}</h1>
    <p><strong>Institution:</strong> {self.config.institution_name}</p>
    <p><strong>Generated:</strong> {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
    <hr>
"""

    def _generate_interactive_dashboard(
        self, components: dict[str, Any], output_path: Path
    ) -> Path:
        """Generate interactive dashboard using Plotly."""
        if not HAS_PLOTLY:
            return None

        # Create comprehensive dashboard combining all visualizations
        fig = make_subplots(
            rows=2,
            cols=2,
            subplot_titles=[
                "Performance Overview",
                "Risk Analysis",
                "Feasibility Assessment",
                "Strategic Recommendations",
            ],
            specs=[[{"type": "xy"}, {"type": "xy"}], [{"type": "xy"}, {"type": "table"}]],
        )

        # Add executive dashboard if available
        if (
            "executive_summary" in components
            and "dashboard_figure" in components["executive_summary"]
        ):
            exec_fig = components["executive_summary"]["dashboard_figure"]
            if exec_fig:
                # Extract data from executive figure and add to subplot
                for trace in exec_fig.data:
                    fig.add_trace(trace, row=1, col=1)

        # Add placeholder content for other subplots
        fig.add_trace(go.Scatter(x=[1, 2, 3], y=[4, 5, 6], name="Risk Metric"), row=1, col=2)
        fig.add_trace(go.Scatter(x=[1, 2, 3], y=[2, 3, 1], name="Feasibility Score"), row=2, col=1)

        # Update layout
        fig.update_layout(
            title_text=f"{self.config.report_title} - Interactive Dashboard",
            showlegend=True,
            height=800,
        )

        # Save interactive dashboard
        fig.write_html(output_path)
        return output_path

    def create_executive_presentation(
        self,
        report_components: dict[str, Any],
        output_path: Path = None,
    ) -> Path:
        """
        Create executive presentation summary from report components.

        Args:
            report_components: Generated report components
            output_path: Optional output path for presentation

        Returns:
            Path to generated presentation file
        """
        if output_path is None:
            output_path = Path(self.config.output_directory) / "executive_presentation.html"

        # Extract key insights for presentation
        key_insights = self._extract_key_insights(report_components)

        # Create presentation HTML
        presentation_html = self._create_presentation_template(key_insights)

        with open(output_path, "w", encoding="utf-8") as f:
            f.write(presentation_html)

        return output_path

    def _extract_key_insights(self, components: dict[str, Any]) -> dict[str, str]:
        """Extract key insights from report components for executive presentation."""
        insights = {}

        if "executive_summary" in components:
            summary_stats = components["executive_summary"].get("summary_statistics", {})
            insights["top_performer"] = summary_stats.get("top_performer", "N/A")
            insights["best_sharpe"] = f"{summary_stats.get('best_sharpe_ratio', 0):.3f}"
            insights["recommendation"] = summary_stats.get("recommended_model", "N/A")

        if "feasibility_assessment" in components:
            insights["implementation_complexity"] = "Medium"  # Simplified for presentation

        if "strategic_recommendations" in components:
            decision_framework = components["strategic_recommendations"].get(
                "decision_framework", {}
            )
            insights["primary_decision"] = decision_framework.get("Primary Decision", "N/A")

        return insights

    def _create_presentation_template(self, insights: dict[str, str]) -> str:
        """Create executive presentation template."""
        return f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Executive Presentation - {self.config.report_title}</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 0;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
        }}
        .slide {{
            width: 100vw;
            height: 100vh;
            display: flex;
            flex-direction: column;
            justify-content: center;
            align-items: center;
            text-align: center;
            padding: 40px;
            box-sizing: border-box;
        }}
        h1 {{
            font-size: 3em;
            margin-bottom: 0.5em;
        }}
        h2 {{
            font-size: 2em;
            margin-bottom: 1em;
            color: #f8f9fa;
        }}
        .insight {{
            font-size: 1.5em;
            margin: 20px 0;
            padding: 20px;
            background: rgba(255,255,255,0.1);
            border-radius: 10px;
            backdrop-filter: blur(10px);
        }}
        .highlight {{
            color: #ffd700;
            font-weight: bold;
        }}
    </style>
</head>
<body>
    <div class="slide">
        <h1>{self.config.report_title}</h1>
        <h2>Executive Summary</h2>

        <div class="insight">
            <strong>Top Performing Model:</strong>
            <span class="highlight">{insights.get('top_performer', 'N/A')}</span>
        </div>

        <div class="insight">
            <strong>Best Sharpe Ratio:</strong>
            <span class="highlight">{insights.get('best_sharpe', 'N/A')}</span>
        </div>

        <div class="insight">
            <strong>Strategic Recommendation:</strong>
            <span class="highlight">{insights.get('recommendation', 'N/A')}</span>
        </div>

        <div class="insight">
            <strong>Primary Decision:</strong><br>
            {insights.get('primary_decision', 'N/A')}
        </div>

        <p style="margin-top: 40px; font-size: 0.9em; opacity: 0.8;">
            Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}
        </p>
    </div>
</body>
</html>
        """
